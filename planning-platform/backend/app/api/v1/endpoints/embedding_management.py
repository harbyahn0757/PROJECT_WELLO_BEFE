"""
병원별 RAG 임베딩 관리 API (백오피스용)

- 파트너/병원 계층 구조
- 병원 목록 및 임베딩 유무
- 병원별 문서 목록·업로드·삭제
- 병원별 인덱스 재구축 트리거
- 병원별 RAG/LLM 설정 CRUD
- 대화 로그 조회
"""

import os
import json
import logging
import traceback
from pathlib import Path
from typing import List, Optional, Dict, Any
from datetime import datetime

from fastapi import APIRouter, HTTPException, UploadFile, File, BackgroundTasks, Body
from pydantic import BaseModel, Field

from ....core.database import db_manager

router = APIRouter(prefix="/embedding", tags=["embedding-management"])

logger = logging.getLogger(__name__)

# 병원별 FAISS 루트 (전역 faiss_db와 분리)
LOCAL_FAISS_BY_HOSPITAL = os.environ.get("LOCAL_FAISS_BY_HOSPITAL", "/data/vector_db/welno/faiss_db_by_hospital")
UPLOADS_SUBDIR = "uploads"

# 임베딩 모델 설정 (rag_service.py와 동일)
EMBEDDING_MODEL = "text-embedding-ada-002"
EMBEDDING_DIMENSION = 1536

# 재구축 진행 상태 추적 (메모리 딕셔너리)
_rebuild_status: Dict[str, Dict[str, Any]] = {}


def _hospital_base_path(hospital_id: str) -> Path:
    return Path(LOCAL_FAISS_BY_HOSPITAL) / hospital_id


def _hospital_uploads_path(hospital_id: str) -> Path:
    return _hospital_base_path(hospital_id) / UPLOADS_SUBDIR

def _common_uploads_path(partner_id: str) -> Path:
    """공통(기본 지침) 문서 경로: _common/{partner_id}/uploads/"""
    return Path(LOCAL_FAISS_BY_HOSPITAL) / "_common" / partner_id / UPLOADS_SUBDIR



# ─── Pydantic 모델 ───────────────────────────────────────────────

class HospitalEmbeddingItem(BaseModel):
    partner_id: str
    partner_name: str
    hospital_id: str
    hospital_name: str
    has_embedding: bool = False
    has_uploads: bool = False
    document_count: int = 0
    chat_count_today: int = 0
    survey_count_today: int = 0


class DocumentItem(BaseModel):
    id: Optional[int] = None
    name: str                          # stored_filename (기존 호환)
    title: str = ""                    # 표시용 제목
    category: Optional[str] = None
    doc_type: str = "hospital"         # global | common | hospital
    size_bytes: int = 0
    chunk_count: int = 0
    uploaded_at: Optional[str] = None
    is_active: bool = True


class HospitalRagConfigUpdate(BaseModel):
    partner_id: str = "welno"
    hospital_name: Optional[str] = None
    contact_phone: Optional[str] = None
    persona_prompt: Optional[str] = None
    welcome_message: Optional[str] = None
    llm_config: Optional[Dict[str, Any]] = None
    embedding_config: Optional[Dict[str, Any]] = None
    theme_config: Optional[Dict[str, Any]] = None
    is_active: Optional[bool] = True


class HospitalRagConfigResponse(BaseModel):
    """병원 RAG 설정 응답 모델"""
    partner_id: str
    hospital_id: str
    hospital_name: Optional[str] = None
    contact_phone: Optional[str] = None
    persona_prompt: Optional[str] = None
    welcome_message: Optional[str] = None
    llm_config: Dict[str, Any]
    embedding_config: Dict[str, Any]
    theme_config: Dict[str, Any]
    is_active: bool
    created_at: Optional[str] = None
    updated_at: Optional[str] = None


class PartnerItem(BaseModel):
    partner_id: str
    partner_name: str
    is_active: bool
    hospital_count: int = 0


class PartnerHierarchy(BaseModel):
    """파트너-병원 계층 구조"""
    partner_id: str
    partner_name: str
    is_active: bool
    hospitals: List[HospitalEmbeddingItem]


# ─── 파트너 / 계층 API ───────────────────────────────────────────

@router.get("/partners", response_model=List[PartnerItem])
async def list_partners():
    """파트너 목록 조회 (병원 수 포함)"""
    try:
        query = """
            SELECT
                p.partner_id,
                p.partner_name,
                p.is_active,
                COUNT(h.hospital_id) as hospital_count
            FROM welno.tb_partner_config p
            LEFT JOIN welno.tb_hospital_rag_config h ON p.partner_id = h.partner_id AND h.hospital_id != '*'
            WHERE p.is_active = true
            GROUP BY p.partner_id, p.partner_name, p.is_active
            ORDER BY p.partner_name
        """
        return await db_manager.execute_query(query)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"파트너 목록 조회 실패: {str(e)}")


@router.get("/summary-counts")
async def get_summary_counts():
    """사이드바 뱃지용 — 오늘 신규 상담 건수 + 설문 건수"""
    try:
        chat_row = await db_manager.execute_one(
            "SELECT COUNT(*) as cnt FROM welno.tb_partner_rag_chat_log WHERE created_at::date = CURRENT_DATE"
        )
        survey_row = await db_manager.execute_one("""
            SELECT
                COALESCE((SELECT COUNT(*) FROM welno.tb_hospital_survey_responses WHERE created_at::date = CURRENT_DATE), 0)
                + COALESCE((SELECT COUNT(*) FROM welno.tb_survey_responses_dynamic WHERE created_at::date = CURRENT_DATE), 0)
            as cnt
        """)
        return {
            "new_chats": chat_row["cnt"] if chat_row else 0,
            "new_surveys": survey_row["cnt"] if survey_row else 0,
        }
    except Exception as e:
        logger.warning(f"summary-counts 조회 실패: {e}")
        return {"new_chats": 0, "new_surveys": 0}


@router.get("/hierarchy", response_model=List[PartnerHierarchy])
async def get_partner_hospital_hierarchy():
    """파트너-병원 계층 구조 조회 (Tree 형태)"""
    try:
        partners_query = """
            SELECT partner_id, partner_name, is_active
            FROM welno.tb_partner_config
            WHERE is_active = true
            ORDER BY partner_name
        """
        partners = await db_manager.execute_query(partners_query)

        result = []
        base = Path(LOCAL_FAISS_BY_HOSPITAL)

        for partner in partners:
            pid = partner['partner_id']
            hospitals_query = """
                SELECT
                    rc.partner_id,
                    COALESCE(pc.partner_name, rc.partner_id) as partner_name,
                    rc.hospital_id,
                    COALESCE(rc.hospital_name, h.hospital_name, rc.hospital_id) as hospital_name,
                    rc.is_active as config_active
                FROM welno.tb_hospital_rag_config rc
                LEFT JOIN welno.welno_hospitals h ON rc.hospital_id = h.hospital_id
                LEFT JOIN welno.tb_partner_config pc ON rc.partner_id = pc.partner_id
                WHERE rc.partner_id = %s AND rc.hospital_id != '*' AND rc.is_active = true
                ORDER BY hospital_name
            """
            hospitals_data = await db_manager.execute_query(hospitals_query, (pid,))

            # 오늘 병원별 채팅/설문 건수 일괄 조회
            chat_counts = {}
            survey_counts = {}
            try:
                chat_rows = await db_manager.execute_query(
                    "SELECT hospital_id, COUNT(*) as cnt FROM welno.tb_partner_rag_chat_log WHERE partner_id = %s AND created_at::date = CURRENT_DATE GROUP BY hospital_id",
                    (pid,)
                )
                chat_counts = {r['hospital_id']: r['cnt'] for r in chat_rows}

                survey_rows = await db_manager.execute_query(
                    """SELECT hospital_id, COUNT(*) as cnt FROM (
                        SELECT hospital_id FROM welno.tb_hospital_survey_responses WHERE partner_id = %s AND created_at::date = CURRENT_DATE
                        UNION ALL
                        SELECT hospital_id FROM welno.tb_survey_responses_dynamic WHERE partner_id = %s AND created_at::date = CURRENT_DATE
                    ) t GROUP BY hospital_id""",
                    (pid, pid)
                )
                survey_counts = {r['hospital_id']: r['cnt'] for r in survey_rows}
            except Exception:
                pass

            hospitals = []
            for h in hospitals_data:
                hid = h['hospital_id']
                base_path = base / hid
                uploads_path = base_path / UPLOADS_SUBDIR
                has_index = (base_path / "faiss.index").exists()
                has_uploads = uploads_path.exists()
                count = len(list(uploads_path.glob("*"))) if has_uploads else 0

                hospitals.append(HospitalEmbeddingItem(
                    partner_id=h['partner_id'],
                    partner_name=h['partner_name'],
                    hospital_id=hid,
                    hospital_name=h['hospital_name'],
                    has_embedding=has_index,
                    has_uploads=has_uploads,
                    document_count=count,
                    chat_count_today=chat_counts.get(hid, 0),
                    survey_count_today=survey_counts.get(hid, 0),
                ))

            result.append(PartnerHierarchy(
                partner_id=partner['partner_id'],
                partner_name=partner['partner_name'],
                is_active=partner['is_active'],
                hospitals=hospitals,
            ))

        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"계층 구조 조회 실패: {str(e)}")


# ─── 프론트엔드 설정 API ──────────────────────────────────────────

@router.get("/config/frontend")
async def get_frontend_config(partner_id: str = "welno", hospital_id: Optional[str] = None):
    """프론트엔드용 동적 설정 제공"""
    try:
        from ....services.dynamic_config_service import dynamic_config

        default_hospital_id = await dynamic_config.get_default_hospital_id(partner_id)
        mediarc_config = await dynamic_config.get_mediarc_config(partner_id)
        metadata = await dynamic_config.get_partner_metadata(partner_id, hospital_id)

        if not metadata or metadata.get("is_not_found"):
            raise HTTPException(status_code=404, detail="등록되지 않은 파트너 또는 병원입니다.")

        return {
            "partner_id": partner_id,
            "hospital_id": hospital_id,
            "partner_name": metadata["partner_name"],
            "phone_number": metadata["phone_number"],
            "default_hospital_id": default_hospital_id,
            "api_key": mediarc_config["api_key"],
            "mediarc_enabled": mediarc_config["enabled"],
            "welcome_message": metadata["welcome_message"],
            "theme": metadata["theme"],
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"프론트엔드 설정 조회 실패: {str(e)}")


# ─── 대기 병원 등록 API ───────────────────────────────────────────

@router.get("/pending-hospitals")
async def get_pending_hospitals():
    """등록 대기 중인 병원 목록"""
    try:
        query = """
            SELECT id, partner_id, hospital_id, first_seen_at, last_seen_at, request_count, status
            FROM welno.tb_pending_hospital_registration
            WHERE status = 'pending'
            ORDER BY last_seen_at DESC
        """
        return await db_manager.execute_query(query)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"대기 병원 목록 조회 실패: {str(e)}")


# ─── 대화 로그 API ────────────────────────────────────────────────

@router.get("/hospitals/{hospital_id}/chats")
async def get_hospital_chats(hospital_id: str, partner_id: str):
    """특정 병원의 대화 세션 목록 조회"""
    try:
        query = """
            SELECT
                l.session_id,
                l.user_uuid,
                l.message_count,
                l.created_at,
                l.updated_at,
                COALESCE(
                    l.initial_data->'patient_info'->>'name',
                    l.client_info->>'patient_name',
                    l.initial_data->>'name',
                    ''
                ) as user_name,
                COALESCE(
                    l.initial_data->'patient_info'->>'contact',
                    l.client_info->>'patient_contact',
                    l.initial_data->>'phone',
                    ''
                ) as user_phone,
                COALESCE(
                    l.initial_data->'patient_info'->>'gender', ''
                ) as user_gender,
                COALESCE(
                    l.initial_data->'health_metrics'->>'checkup_date', ''
                ) as checkup_date,
                COALESCE(h.hospital_name, '') as hospital_name
            FROM welno.tb_partner_rag_chat_log l
            LEFT JOIN welno.tb_hospital_rag_config h
                ON l.hospital_id = h.hospital_id AND l.partner_id = h.partner_id
            WHERE l.partner_id = %s AND l.hospital_id = %s
            ORDER BY l.created_at DESC
        """
        return await db_manager.execute_query(query, (partner_id, hospital_id))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"대화 목록 조회 실패: {str(e)}")


@router.get("/chats/all")
async def get_all_chats(partner_id: Optional[str] = None, limit: int = 200):
    """전체 병원 통합 대화 세션 목록 조회 (태그 포함)"""
    try:
        if partner_id:
            query = """
                SELECT
                    l.session_id,
                    l.partner_id,
                    l.hospital_id,
                    l.user_uuid,
                    l.message_count,
                    l.created_at,
                    l.updated_at,
                    COALESCE(
                        l.initial_data->'patient_info'->>'name',
                        l.client_info->>'patient_name',
                        ''
                    ) as user_name,
                    COALESCE(
                        l.initial_data->'patient_info'->>'contact',
                        l.client_info->>'patient_contact',
                        ''
                    ) as user_phone,
                    COALESCE(
                        l.initial_data->'patient_info'->>'gender', ''
                    ) as user_gender,
                    COALESCE(
                        l.initial_data->'health_metrics'->>'checkup_date', ''
                    ) as checkup_date,
                    COALESCE(h.hospital_name, LEFT(l.hospital_id, 8) || '...') as hospital_name,
                    t.interest_tags,
                    t.risk_tags,
                    t.sentiment,
                    t.conversation_summary,
                    t.data_quality_score
                FROM welno.tb_partner_rag_chat_log l
                LEFT JOIN welno.tb_hospital_rag_config h
                    ON l.hospital_id = h.hospital_id AND l.partner_id = h.partner_id
                LEFT JOIN welno.tb_chat_session_tags t
                    ON l.session_id = t.session_id AND l.partner_id = t.partner_id
                WHERE l.partner_id = %s
                ORDER BY l.created_at DESC
                LIMIT %s
            """
            return await db_manager.execute_query(query, (partner_id, limit))
        else:
            query = """
                SELECT
                    l.session_id,
                    l.partner_id,
                    l.hospital_id,
                    l.user_uuid,
                    l.message_count,
                    l.created_at,
                    l.updated_at,
                    COALESCE(
                        l.initial_data->'patient_info'->>'name',
                        l.client_info->>'patient_name',
                        ''
                    ) as user_name,
                    COALESCE(
                        l.initial_data->'patient_info'->>'contact',
                        l.client_info->>'patient_contact',
                        ''
                    ) as user_phone,
                    COALESCE(
                        l.initial_data->'patient_info'->>'gender', ''
                    ) as user_gender,
                    COALESCE(
                        l.initial_data->'health_metrics'->>'checkup_date', ''
                    ) as checkup_date,
                    COALESCE(h.hospital_name, LEFT(l.hospital_id, 8) || '...') as hospital_name,
                    t.interest_tags,
                    t.risk_tags,
                    t.sentiment,
                    t.conversation_summary,
                    t.data_quality_score
                FROM welno.tb_partner_rag_chat_log l
                LEFT JOIN welno.tb_hospital_rag_config h
                    ON l.hospital_id = h.hospital_id AND l.partner_id = h.partner_id
                LEFT JOIN welno.tb_chat_session_tags t
                    ON l.session_id = t.session_id AND l.partner_id = t.partner_id
                ORDER BY l.created_at DESC
                LIMIT %s
            """
            return await db_manager.execute_query(query, (limit,))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"통합 대화 목록 조회 실패: {str(e)}")


@router.get("/chats/export")
async def export_chats_excel(partner_id: Optional[str] = None, limit: int = 500):
    """대화 세션 목록을 엑셀(XLSX)로 내보내기"""
    from fastapi.responses import StreamingResponse
    import io

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 라이브러리가 설치되지 않았습니다.")

    try:
        # 1. 세션 목록 조회 (태그 포함)
        if partner_id:
            query = """
                SELECT
                    l.session_id, l.partner_id, l.hospital_id, l.user_uuid,
                    l.message_count, l.created_at, l.updated_at,
                    l.conversation, l.initial_data,
                    COALESCE(l.initial_data->'patient_info'->>'name', l.client_info->>'patient_name', '') as user_name,
                    COALESCE(l.initial_data->'patient_info'->>'contact', l.client_info->>'patient_contact', '') as user_phone,
                    COALESCE(l.initial_data->'patient_info'->>'gender', '') as user_gender,
                    COALESCE(l.initial_data->'health_metrics'->>'checkup_date', '') as checkup_date,
                    COALESCE(h.hospital_name, '') as hospital_name,
                    t.interest_tags, t.risk_tags, t.sentiment,
                    t.conversation_summary, t.data_quality_score,
                    t.risk_level, t.key_concerns, t.follow_up_needed,
                    t.counselor_recommendations, t.keyword_tags,
                    t.conversation_depth, t.engagement_score, t.action_intent, t.nutrition_tags
                FROM welno.tb_partner_rag_chat_log l
                LEFT JOIN welno.tb_hospital_rag_config h ON l.hospital_id = h.hospital_id AND l.partner_id = h.partner_id
                LEFT JOIN welno.tb_chat_session_tags t ON l.session_id = t.session_id AND l.partner_id = t.partner_id
                WHERE l.partner_id = %s
                ORDER BY l.created_at DESC LIMIT %s
            """
            rows = await db_manager.execute_query(query, (partner_id, limit))
        else:
            query = """
                SELECT
                    l.session_id, l.partner_id, l.hospital_id, l.user_uuid,
                    l.message_count, l.created_at, l.updated_at,
                    l.conversation, l.initial_data,
                    COALESCE(l.initial_data->'patient_info'->>'name', l.client_info->>'patient_name', '') as user_name,
                    COALESCE(l.initial_data->'patient_info'->>'contact', l.client_info->>'patient_contact', '') as user_phone,
                    COALESCE(l.initial_data->'patient_info'->>'gender', '') as user_gender,
                    COALESCE(l.initial_data->'health_metrics'->>'checkup_date', '') as checkup_date,
                    COALESCE(h.hospital_name, '') as hospital_name,
                    t.interest_tags, t.risk_tags, t.sentiment,
                    t.conversation_summary, t.data_quality_score,
                    t.risk_level, t.key_concerns, t.follow_up_needed,
                    t.counselor_recommendations, t.keyword_tags,
                    t.conversation_depth, t.engagement_score, t.action_intent, t.nutrition_tags
                FROM welno.tb_partner_rag_chat_log l
                LEFT JOIN welno.tb_hospital_rag_config h ON l.hospital_id = h.hospital_id AND l.partner_id = h.partner_id
                LEFT JOIN welno.tb_chat_session_tags t ON l.session_id = t.session_id AND l.partner_id = t.partner_id
                ORDER BY l.created_at DESC LIMIT %s
            """
            rows = await db_manager.execute_query(query, (limit,))

        # 2. 워크북 생성 (단일 시트)
        wb = openpyxl.Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        wrap_align = Alignment(wrap_text=True, vertical="top")

        ws = wb.active
        ws.title = "상담 데이터"
        headers = [
            "세션ID", "파트너", "병원", "환자명", "성별", "연락처", "검진일",
            "메시지수", "관심사 태그", "위험 태그", "위험도", "감정",
            "대화깊이", "참여도점수", "행동의향", "식단·영양 관심",
            "핵심 우려사항", "후속조치 필요", "상담사 권고사항",
            "요약", "키워드 태그", "데이터품질",
            "생성일", "검진 데이터(JSON)", "대화 내역(JSON)"
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, r in enumerate(rows, 2):
            # interest_tags: [{topic, intensity}] → "혈압(high), 당뇨(medium)" 형식
            raw_interest = r.get("interest_tags") or []
            if isinstance(raw_interest, str):
                try:
                    raw_interest = json.loads(raw_interest)
                except:
                    raw_interest = []
            if isinstance(raw_interest, list):
                interest_parts = []
                for item in raw_interest:
                    if isinstance(item, dict) and "topic" in item:
                        intensity = item.get("intensity", "medium")
                        interest_parts.append(f"{item['topic']}({intensity})")
                    elif isinstance(item, str):
                        interest_parts.append(item)
                interest = ", ".join(interest_parts)
            else:
                interest = ""

            risk = ", ".join(r.get("risk_tags") or []) if r.get("risk_tags") else ""

            # key_concerns
            raw_concerns = r.get("key_concerns") or []
            if isinstance(raw_concerns, str):
                try:
                    raw_concerns = json.loads(raw_concerns)
                except:
                    raw_concerns = []
            concerns_str = ", ".join(raw_concerns) if isinstance(raw_concerns, list) else ""

            # counselor_recommendations
            raw_recs = r.get("counselor_recommendations") or []
            if isinstance(raw_recs, str):
                try:
                    raw_recs = json.loads(raw_recs)
                except:
                    raw_recs = []
            recs_str = ", ".join(raw_recs) if isinstance(raw_recs, list) else ""

            # keyword_tags
            raw_kw = r.get("keyword_tags") or []
            if isinstance(raw_kw, str):
                try:
                    raw_kw = json.loads(raw_kw)
                except:
                    raw_kw = []
            kw_str = ", ".join(raw_kw) if isinstance(raw_kw, list) else ""

            # nutrition_tags
            raw_nutrition = r.get("nutrition_tags") or []
            if isinstance(raw_nutrition, str):
                try:
                    raw_nutrition = json.loads(raw_nutrition)
                except:
                    raw_nutrition = []
            nutrition_str = ", ".join(raw_nutrition) if isinstance(raw_nutrition, list) else ""

            # 검진 데이터 JSON
            initial = r.get("initial_data") or {}
            if isinstance(initial, str):
                try:
                    initial = json.loads(initial)
                except:
                    initial = {}
            metrics = initial.get("health_metrics", {})
            metrics_json = json.dumps(metrics, ensure_ascii=False) if metrics else ""

            # 대화 내역 JSON
            conversation = r.get("conversation") or []
            if isinstance(conversation, str):
                try:
                    conversation = json.loads(conversation)
                except:
                    conversation = []
            conv_json = json.dumps(conversation, ensure_ascii=False) if conversation else ""
            # Excel 셀 최대 32767자 제한
            if len(conv_json) > 32000:
                conv_json = conv_json[:32000] + "...(truncated)"

            col = 1
            ws.cell(row=row_idx, column=col, value=r.get("session_id", "")); col += 1
            ws.cell(row=row_idx, column=col, value=r.get("partner_id", "")); col += 1
            ws.cell(row=row_idx, column=col, value=r.get("hospital_name", "")); col += 1
            ws.cell(row=row_idx, column=col, value=r.get("user_name", "")); col += 1
            ws.cell(row=row_idx, column=col, value=r.get("user_gender", "")); col += 1
            ws.cell(row=row_idx, column=col, value=r.get("user_phone", "")); col += 1
            ws.cell(row=row_idx, column=col, value=r.get("checkup_date", "")); col += 1
            ws.cell(row=row_idx, column=col, value=r.get("message_count", 0)); col += 1
            ws.cell(row=row_idx, column=col, value=interest); col += 1
            ws.cell(row=row_idx, column=col, value=risk); col += 1
            ws.cell(row=row_idx, column=col, value=r.get("risk_level", "")); col += 1
            ws.cell(row=row_idx, column=col, value=r.get("sentiment", "")); col += 1
            ws.cell(row=row_idx, column=col, value=r.get("conversation_depth", "")); col += 1
            ws.cell(row=row_idx, column=col, value=r.get("engagement_score") or 0); col += 1
            ws.cell(row=row_idx, column=col, value=r.get("action_intent", "")); col += 1
            ws.cell(row=row_idx, column=col, value=nutrition_str); col += 1
            ws.cell(row=row_idx, column=col, value=concerns_str); col += 1
            ws.cell(row=row_idx, column=col, value="Y" if r.get("follow_up_needed") else "N"); col += 1
            ws.cell(row=row_idx, column=col, value=recs_str); col += 1
            ws.cell(row=row_idx, column=col, value=r.get("conversation_summary", "")); col += 1
            ws.cell(row=row_idx, column=col, value=kw_str); col += 1
            ws.cell(row=row_idx, column=col, value=r.get("data_quality_score") or 0); col += 1
            ws.cell(row=row_idx, column=col, value=str(r.get("created_at", ""))); col += 1
            cell_metrics = ws.cell(row=row_idx, column=col, value=metrics_json); col += 1
            cell_metrics.alignment = wrap_align
            cell_conv = ws.cell(row=row_idx, column=col, value=conv_json)
            cell_conv.alignment = wrap_align

        # 3. 버퍼에 저장 후 반환
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"welno_chats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"엑셀 내보내기 실패: {e}")
        raise HTTPException(status_code=500, detail=f"엑셀 내보내기 실패: {str(e)}")


@router.get("/chats/{session_id}")
async def get_chat_detail(session_id: str):
    """특정 세션의 대화 상세 내역 조회 (태그 포함)"""
    try:
        query = """
            SELECT
                l.session_id,
                l.partner_id,
                l.hospital_id,
                l.user_uuid,
                l.conversation,
                l.initial_data,
                l.created_at,
                t.interest_tags,
                t.risk_tags,
                t.keyword_tags,
                t.sentiment,
                t.conversation_summary,
                t.data_quality_score,
                t.has_discrepancy,
                t.risk_level,
                t.key_concerns,
                t.follow_up_needed,
                t.tagging_model,
                t.tagging_version,
                t.counselor_recommendations
            FROM welno.tb_partner_rag_chat_log l
            LEFT JOIN welno.tb_chat_session_tags t
                ON l.session_id = t.session_id AND l.partner_id = t.partner_id
            WHERE l.session_id = %s
        """
        result = await db_manager.execute_one(query, (session_id,))
        if not result:
            raise HTTPException(status_code=404, detail="대화 내역을 찾을 수 없습니다.")
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"대화 상세 조회 실패: {str(e)}")


# ─── 재태깅 API ──────────────────────────────────────────────────

class RetagRequest(BaseModel):
    hospital_id: Optional[str] = None
    force: bool = False

@router.post("/retag-sessions")
async def retag_sessions(body: RetagRequest = Body(...)):
    """기존 세션 일괄 재태깅 (LLM 기반). force=true면 이미 LLM 태깅된 것도 재처리."""
    try:
        from ....services.chat_tagging_service import retag_all_sessions
        result = await retag_all_sessions(
            hospital_id=body.hospital_id,
            force=body.force,
        )
        return {"success": True, **result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"재태깅 실패: {str(e)}")

@router.post("/chats/{session_id}/retag")
async def retag_single_session(session_id: str):
    """단일 세션 재태깅 (LLM 기반)"""
    try:
        from ....services.chat_tagging_service import tag_chat_session
        # 세션의 partner_id 조회
        query = "SELECT partner_id FROM welno.tb_partner_rag_chat_log WHERE session_id = %s"
        row = await db_manager.execute_one(query, (session_id,))
        if not row:
            raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다.")
        result = await tag_chat_session(session_id=session_id, partner_id=row["partner_id"])
        if result:
            return {"success": True, "session_id": session_id, "tagging_model": result.get("tagging_model")}
        raise HTTPException(status_code=500, detail="태깅 결과 없음")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"재태깅 실패: {str(e)}")


# ─── 병원 목록 API ────────────────────────────────────────────────

@router.get("/hospitals", response_model=List[HospitalEmbeddingItem])
async def list_hospitals_with_embedding_status(partner_id: Optional[str] = None):
    """병원 목록 조회 (파트너 계층 포함, 설정 테이블 기준)"""
    try:
        where_clause = "WHERE rc.hospital_id != '*'"
        params = []

        if partner_id:
            where_clause += " AND rc.partner_id = %s"
            params.append(partner_id)

        query = f"""
            SELECT
                rc.partner_id,
                COALESCE(pc.partner_name, rc.partner_id) as partner_name,
                rc.hospital_id,
                COALESCE(rc.hospital_name, h.hospital_name, rc.hospital_id) as hospital_name,
                rc.is_active as config_active
            FROM welno.tb_hospital_rag_config rc
            LEFT JOIN welno.welno_hospitals h ON rc.hospital_id = h.hospital_id
            LEFT JOIN welno.tb_partner_config pc ON rc.partner_id = pc.partner_id
            {where_clause}
            ORDER BY partner_name, hospital_name
        """
        mappings = await db_manager.execute_query(query, params)

        remaining = []
        if not partner_id:
            query_remaining = """
                SELECT
                    'welno' as partner_id,
                    'Welno (기본)' as partner_name,
                    h.hospital_id,
                    h.hospital_name,
                    true as config_active
                FROM welno.welno_hospitals h
                WHERE h.hospital_id NOT IN (SELECT hospital_id FROM welno.tb_hospital_rag_config WHERE hospital_id != '*')
                AND h.is_active = true
            """
            remaining = await db_manager.execute_query(query_remaining)

        all_hospitals = mappings + remaining

        result = []
        base = Path(LOCAL_FAISS_BY_HOSPITAL)

        for m in all_hospitals:
            hid = m['hospital_id']
            base_path = base / hid
            uploads_path = base_path / UPLOADS_SUBDIR
            has_index = (base_path / "faiss.index").exists()
            has_uploads = uploads_path.exists()
            count = len(list(uploads_path.glob("*"))) if has_uploads else 0

            result.append(HospitalEmbeddingItem(
                partner_id=m['partner_id'],
                partner_name=m['partner_name'],
                hospital_id=hid,
                hospital_name=m['hospital_name'],
                has_embedding=has_index,
                has_uploads=has_uploads,
                document_count=count,
            ))
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"병원 목록 조회 실패: {str(e)}")


# ─── 문서 관리 API ────────────────────────────────────────────────

@router.get("/hospitals/{hospital_id}/documents", response_model=List[DocumentItem])
async def list_hospital_documents(hospital_id: str):
    """병원별 업로드 문서 목록 (DB 조회)"""
    try:
        query = """
            SELECT id, title, category, doc_type, original_filename, stored_filename,
                   size_bytes, chunk_count, created_at, is_active
            FROM welno.tb_rag_documents
            WHERE hospital_id = %s AND doc_type = 'hospital'
            ORDER BY is_active DESC, created_at DESC
        """
        rows = await db_manager.execute_query(query, (hospital_id,))
        items = []
        for r in rows:
            items.append(DocumentItem(
                id=r["id"],
                name=r["stored_filename"],
                title=r.get("title") or r["stored_filename"],
                category=r.get("category"),
                doc_type=r["doc_type"],
                size_bytes=r.get("size_bytes") or 0,
                chunk_count=r.get("chunk_count") or 0,
                uploaded_at=r["created_at"].isoformat() if r.get("created_at") else None,
                is_active=r.get("is_active", True),
            ))
        # DB에 없으면 파일시스템 폴백 (마이그레이션 전 호환)
        if not items:
            uploads_path = _hospital_uploads_path(hospital_id)
            if uploads_path.exists():
                for f in uploads_path.iterdir():
                    if f.is_file() and not f.name.startswith("."):
                        stat = f.stat()
                        items.append(DocumentItem(
                            name=f.name,
                            title=os.path.splitext(f.name)[0].replace("_", " "),
                            size_bytes=stat.st_size,
                            uploaded_at=datetime.fromtimestamp(stat.st_mtime).isoformat(),
                        ))
                items.sort(key=lambda x: x.uploaded_at or "", reverse=True)
        return items
    except Exception as e:
        logger.error(f"문서 목록 조회 실패: {e}")
        # DB 실패 시 파일시스템 폴백
        uploads_path = _hospital_uploads_path(hospital_id)
        if not uploads_path.exists():
            return []
        items = []
        for f in uploads_path.iterdir():
            if f.is_file() and not f.name.startswith("."):
                stat = f.stat()
                items.append(DocumentItem(
                    name=f.name,
                    title=os.path.splitext(f.name)[0].replace("_", " "),
                    size_bytes=stat.st_size,
                    uploaded_at=datetime.fromtimestamp(stat.st_mtime).isoformat(),
                ))
        items.sort(key=lambda x: x.uploaded_at or "", reverse=True)
        return items


@router.post("/hospitals/{hospital_id}/documents")
async def upload_hospital_document(
    hospital_id: str,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    partner_id: str = "welno",
) -> Dict[str, Any]:
    """병원별 문서 업로드 (PDF 등). 업로드 후 자동으로 인덱싱."""
    uploads_path = _hospital_uploads_path(hospital_id)
    uploads_path.mkdir(parents=True, exist_ok=True)
    safe_name = "".join(c for c in file.filename or "upload" if c.isalnum() or c in "._- ").strip() or "upload"
    dest = uploads_path / safe_name
    try:
        file_content = await file.read()
        dest.write_bytes(file_content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"저장 실패: {str(e)}")
    # DB에 레코드 생성
    title = os.path.splitext(safe_name)[0].replace("_", " ")
    doc_id = None
    try:
        insert_query = """
            INSERT INTO welno.tb_rag_documents
            (partner_id, hospital_id, doc_type, title, original_filename, stored_filename, file_path, size_bytes)
            VALUES (%s, %s, 'hospital', %s, %s, %s, %s, %s)
            RETURNING id
        """
        result = await db_manager.execute_one(insert_query, (
            partner_id, hospital_id, title,
            file.filename or safe_name, safe_name, str(dest), len(file_content)
        ))
        doc_id = result["id"] if result else None
    except Exception as db_err:
        logger.warning(f"문서 DB 레코드 생성 실패 (파일은 저장됨): {db_err}")
    # 업로드 후 자동 인덱싱 트리거
    background_tasks.add_task(_run_rebuild_for_hospital, hospital_id, partner_id)
    return {
        "success": True,
        "hospital_id": hospital_id,
        "filename": safe_name,
        "title": title,
        "id": doc_id,
        "size_bytes": len(file_content),
        "indexing": "started",
    }


@router.delete("/hospitals/{hospital_id}/documents/{filename}")
async def delete_hospital_document(hospital_id: str, filename: str) -> Dict[str, Any]:
    """병원별 업로드 문서 비활성화 (소프트 삭제). 파일과 DB 레코드는 보존."""
    try:
        await db_manager.execute_update(
            """UPDATE welno.tb_rag_documents
               SET is_active = false, deleted_at = NOW(), updated_at = NOW()
               WHERE hospital_id = %s AND stored_filename = %s AND doc_type = 'hospital'""",
            (hospital_id, filename)
        )
    except Exception as db_err:
        logger.warning(f"문서 비활성화 실패: {db_err}")
        raise HTTPException(status_code=500, detail=f"비활성화 실패: {str(db_err)}")
    return {
        "success": True,
        "hospital_id": hospital_id,
        "filename": filename,
        "action": "deactivated",
    }



# ─── 공통(기본 지침) 문서 API ──────────────────────────────────

@router.get("/partners/{partner_id}/common-documents", response_model=List[DocumentItem])
async def list_common_documents(partner_id: str):
    """파트너 공통(기본 지침) 문서 + 글로벌 기본문서 목록 (DB 조회)"""
    try:
        query = """
            SELECT id, title, category, doc_type, original_filename, stored_filename,
                   size_bytes, chunk_count, created_at, is_active
            FROM welno.tb_rag_documents
            WHERE ((partner_id = %s AND doc_type IN ('global', 'common'))
               OR (doc_type = 'global'))
               AND is_active = true
            ORDER BY doc_type, category, title
        """
        rows = await db_manager.execute_query(query, (partner_id,))
        # 중복 제거 (global 문서가 2번 매칭될 수 있음)
        seen_ids = set()
        items = []
        for r in rows:
            if r["id"] in seen_ids:
                continue
            seen_ids.add(r["id"])
            items.append(DocumentItem(
                id=r["id"],
                name=r["stored_filename"],
                title=r.get("title") or r["stored_filename"],
                category=r.get("category"),
                doc_type=r["doc_type"],
                size_bytes=r.get("size_bytes") or 0,
                chunk_count=r.get("chunk_count") or 0,
                uploaded_at=r["created_at"].isoformat() if r.get("created_at") else None,
                is_active=r.get("is_active", True),
            ))
        # DB에 없으면 파일시스템 폴백
        if not items:
            uploads_path = _common_uploads_path(partner_id)
            if uploads_path.exists():
                for f in uploads_path.iterdir():
                    if f.is_file() and not f.name.startswith("."):
                        stat = f.stat()
                        items.append(DocumentItem(
                            name=f.name,
                            title=os.path.splitext(f.name)[0].replace("_", " "),
                            doc_type="common",
                            size_bytes=stat.st_size,
                            uploaded_at=datetime.fromtimestamp(stat.st_mtime).isoformat(),
                        ))
                items.sort(key=lambda x: x.uploaded_at or "", reverse=True)
        return items
    except Exception as e:
        logger.error(f"공통 문서 목록 조회 실패: {e}")
        # DB 실패 시 파일시스템 폴백
        uploads_path = _common_uploads_path(partner_id)
        if not uploads_path.exists():
            return []
        items = []
        for f in uploads_path.iterdir():
            if f.is_file() and not f.name.startswith("."):
                stat = f.stat()
                items.append(DocumentItem(
                    name=f.name,
                    title=os.path.splitext(f.name)[0].replace("_", " "),
                    doc_type="common",
                    size_bytes=stat.st_size,
                    uploaded_at=datetime.fromtimestamp(stat.st_mtime).isoformat(),
                ))
        items.sort(key=lambda x: x.uploaded_at or "", reverse=True)
        return items


@router.post("/partners/{partner_id}/common-documents")
async def upload_common_document(
    partner_id: str,
    file: UploadFile = File(...),
) -> Dict[str, Any]:
    """파트너 공통(기본 지침) 문서 업로드."""
    uploads_path = _common_uploads_path(partner_id)
    uploads_path.mkdir(parents=True, exist_ok=True)
    safe_name = "".join(c for c in file.filename or "upload" if c.isalnum() or c in "._- ").strip() or "upload"
    dest = uploads_path / safe_name
    try:
        file_content = await file.read()
        dest.write_bytes(file_content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"저장 실패: {str(e)}")
    # DB에 레코드 생성
    title = os.path.splitext(safe_name)[0].replace("_", " ")
    doc_id = None
    try:
        insert_query = """
            INSERT INTO welno.tb_rag_documents
            (partner_id, hospital_id, doc_type, title, original_filename, stored_filename, file_path, size_bytes)
            VALUES (%s, NULL, 'common', %s, %s, %s, %s, %s)
            RETURNING id
        """
        result = await db_manager.execute_one(insert_query, (
            partner_id, title,
            file.filename or safe_name, safe_name, str(dest), len(file_content)
        ))
        doc_id = result["id"] if result else None
    except Exception as db_err:
        logger.warning(f"공통 문서 DB 레코드 생성 실패 (파일은 저장됨): {db_err}")
    return {
        "success": True,
        "partner_id": partner_id,
        "filename": safe_name,
        "title": title,
        "id": doc_id,
        "size_bytes": len(file_content),
    }


@router.delete("/partners/{partner_id}/common-documents/{filename}")
async def delete_common_document(partner_id: str, filename: str) -> Dict[str, Any]:
    """파트너 공통 문서 비활성화 (소프트 삭제). 파일과 DB 레코드는 보존."""
    try:
        await db_manager.execute_update(
            """UPDATE welno.tb_rag_documents
               SET is_active = false, deleted_at = NOW(), updated_at = NOW()
               WHERE partner_id = %s AND stored_filename = %s AND doc_type = 'common'""",
            (partner_id, filename)
        )
    except Exception as db_err:
        logger.warning(f"공통 문서 비활성화 실패: {db_err}")
        raise HTTPException(status_code=500, detail=f"비활성화 실패: {str(db_err)}")
    return {
        "success": True,
        "partner_id": partner_id,
        "filename": filename,
        "action": "deactivated",
    }

# ─── 문서 다운로드/미리보기 API ───────────────────────────────────

@router.get("/documents/{doc_id}/download")
async def download_document(doc_id: int):
    """문서 파일 다운로드 (미리보기 겸용)"""
    from fastapi.responses import FileResponse
    try:
        result = await db_manager.execute_one(
            "SELECT file_path, original_filename, stored_filename FROM welno.tb_rag_documents WHERE id = %s",
            (doc_id,)
        )
        if not result:
            raise HTTPException(status_code=404, detail="문서를 찾을 수 없습니다.")
        file_path = Path(result["file_path"])
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="파일이 서버에 없습니다.")
        filename = result.get("original_filename") or result["stored_filename"]
        ext = file_path.suffix.lower()
        media_types = {
            ".pdf": "application/pdf",
            ".txt": "text/plain; charset=utf-8",
            ".md": "text/markdown; charset=utf-8",
            ".csv": "text/csv; charset=utf-8",
        }
        media_type = media_types.get(ext, "application/octet-stream")
        return FileResponse(path=str(file_path), filename=filename, media_type=media_type)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"다운로드 실패: {str(e)}")


@router.put("/documents/{doc_id}/toggle-active")
async def toggle_document_active(doc_id: int, background_tasks: BackgroundTasks) -> Dict[str, Any]:
    """문서 활성/비활성 토글 (FAISS 자동 리빌드 포함)"""
    try:
        current = await db_manager.execute_one(
            "SELECT is_active, hospital_id, partner_id FROM welno.tb_rag_documents WHERE id = %s", (doc_id,)
        )
        if not current:
            raise HTTPException(status_code=404, detail="문서를 찾을 수 없습니다.")
        new_active = not current["is_active"]
        await db_manager.execute_update(
            """UPDATE welno.tb_rag_documents
               SET is_active = %s, deleted_at = CASE WHEN %s = false THEN NOW() ELSE NULL END, updated_at = NOW()
               WHERE id = %s""",
            (new_active, new_active, doc_id)
        )
        # 비활성화/활성화 후 FAISS 인덱스 자동 리빌드
        hospital_id = current.get("hospital_id")
        partner_id = current.get("partner_id", "welno")
        if hospital_id:
            logger.info(f"[토글] 문서 {doc_id} is_active={new_active} -> FAISS 리빌드 트리거 (hospital={hospital_id})")
            background_tasks.add_task(_run_rebuild_for_hospital, hospital_id, partner_id)
        return {"success": True, "id": doc_id, "is_active": new_active, "rebuilding": bool(hospital_id)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/documents/{doc_id}/title")
async def update_document_title(doc_id: int, body: Dict[str, Any]) -> Dict[str, Any]:
    """문서 제목 수정"""
    new_title = body.get("title", "").strip()
    if not new_title:
        raise HTTPException(status_code=400, detail="제목을 입력하세요.")
    try:
        await db_manager.execute_update(
            "UPDATE welno.tb_rag_documents SET title = %s, updated_at = NOW() WHERE id = %s",
            (new_title[:500], doc_id)
        )
        return {"success": True, "id": doc_id, "title": new_title[:500]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── 아이콘 업로드 API ────────────────────────────────────────────

@router.post("/hospitals/{hospital_id}/icon")
async def upload_hospital_icon(
    hospital_id: str,
    file: UploadFile = File(...),
) -> Dict[str, Any]:
    """병원/파트너별 커스텀 아이콘 업로드"""
    try:
        icon_dir = Path("static/uploads/icons")
        icon_dir.mkdir(parents=True, exist_ok=True)

        ext = Path(file.filename).suffix.lower()
        if ext not in [".png", ".jpg", ".jpeg", ".svg", ".gif", ".webp"]:
            raise HTTPException(status_code=400, detail="지원하지 않는 이미지 형식입니다.")

        filename = f"{hospital_id}_icon{ext}"
        dest = icon_dir / filename

        content = await file.read()
        dest.write_bytes(content)

        url = f"/static/uploads/icons/{filename}"
        return {"success": True, "url": url}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"아이콘 업로드 실패: {str(e)}")


# ─── 임베딩 재구축 API ────────────────────────────────────────────

def _run_rebuild_for_hospital(hospital_id: str, partner_id: str = "welno") -> None:
    """병원별 FAISS 인덱스 재구축 (백그라운드).

    uploads 폴더의 PDF/텍스트를 파싱 → OpenAI 임베딩 → FAISS 인덱스 저장.
    rag_service.py와 동일한 EMBEDDING_MODEL, EMBEDDING_DIMENSION 사용.
    """
    global _rebuild_status
    _rebuild_status[hospital_id] = {
        "status": "running",
        "started_at": datetime.now().isoformat(),
        "progress": "초기화 중...",
        "error": None,
    }

    base = _hospital_base_path(hospital_id)
    uploads = base / UPLOADS_SUBDIR
    base.mkdir(parents=True, exist_ok=True)

    common_uploads = _common_uploads_path(partner_id)
    
    has_hospital_docs = uploads.exists() and any(uploads.iterdir())
    has_common_docs = common_uploads.exists() and any(common_uploads.iterdir())
    
    if not has_hospital_docs and not has_common_docs:
        _rebuild_status[hospital_id].update({
            "status": "failed",
            "error": "업로드된 문서가 없습니다.",
            "finished_at": datetime.now().isoformat(),
        })
        return

    try:
        # 1. 필수 라이브러리 임포트
        from llama_index.core import (
            Settings,
            VectorStoreIndex,
            StorageContext,
            Document,
        )
        from llama_index.embeddings.openai import OpenAIEmbedding
        from llama_index.vector_stores.faiss import FaissVectorStore
        import faiss
        from pypdf import PdfReader

        _rebuild_status[hospital_id]["progress"] = "문서 파싱 중..."

        # 2. 공통 + 병원 uploads 내 파일 → LlamaIndex Document 변환
        documents: list = []
        supported_exts = {".pdf", ".txt", ".md", ".csv"}
        
        # 비활성 문서 목록 (인덱싱 제외)
        inactive_files = set()
        try:
            import psycopg2 as _pg
            _conn = _pg.connect(host="10.0.1.10", port=5432, dbname="p9_mkt_biz", user="peernine", password="autumn3334!")
            _cur = _conn.cursor()
            _cur.execute("SELECT stored_filename FROM welno.tb_rag_documents WHERE is_active = false AND (hospital_id = %s OR doc_type = 'common')", (hospital_id,))
            for row in _cur.fetchall():
                inactive_files.add(row[0])
            _cur.close()
            _conn.close()
            if inactive_files:
                logger.info(f"[rebuild:{hospital_id}] 비활성 문서 {len(inactive_files)}개 제외")
        except Exception as _db_err:
            logger.warning(f"[rebuild:{hospital_id}] 비활성 문서 조회 실패: {_db_err}")

        # 공통/병원 문서 순회 목록 (공통 먼저)
        doc_sources = []
        if has_common_docs:
            doc_sources.append(("common", common_uploads))
        if has_hospital_docs:
            doc_sources.append(("hospital", uploads))
        
        for source_type, source_dir in doc_sources:
            for filepath in sorted(source_dir.iterdir()):
                if not filepath.is_file() or filepath.name.startswith("."):
                    continue
                if filepath.name in inactive_files:
                    logger.info(f"[rebuild:{hospital_id}] 건너뜀 (비활성): {filepath.name}")
                    continue
                ext = filepath.suffix.lower()
                if ext not in supported_exts:
                    logger.info(f"[rebuild:{hospital_id}] 건너뜀 (미지원): {filepath.name}")
                    continue
                try:
                    if ext == ".pdf":
                        reader = PdfReader(str(filepath))
                        for page_num, page in enumerate(reader.pages, 1):
                            text = page.extract_text() or ""
                            text = text.strip()
                            if text:
                                documents.append(Document(
                                    text=text,
                                    metadata={
                                        "file_name": filepath.name,
                                        "page_label": str(page_num),
                                        "hospital_id": hospital_id,
                                        "source_type": source_type,
                                    },
                                ))
                    else:
                        # txt, md, csv → 전체 텍스트
                        text = filepath.read_text(encoding="utf-8", errors="ignore").strip()
                        if text:
                            documents.append(Document(
                                text=text,
                                metadata={
                                    "file_name": filepath.name,
                                    "page_label": "1",
                                    "hospital_id": hospital_id,
                                    "source_type": source_type,
                                },
                            ))
                except Exception as parse_err:
                    logger.warning(f"[rebuild:{hospital_id}] 파싱 실패 {filepath.name}: {parse_err}")

        if not documents:
            # 활성 문서 0개 -> 기존 FAISS 인덱스 삭제 (비활성 문서가 검색되지 않도록)
            for old_file in ["faiss.index", "default__vector_store.json", "docstore.json",
                             "index_store.json", "graph_store.json", "image__vector_store.json"]:
                old_path = base / old_file
                if old_path.exists():
                    old_path.unlink()
                    logger.info(f"[rebuild:{hospital_id}] 기존 인덱스 삭제: {old_file}")
            _rebuild_status[hospital_id].update({
                "status": "completed",
                "progress": "활성 문서 없음 - 인덱스 초기화 완료",
                "finished_at": datetime.now().isoformat(),
                "document_count": 0,
                "vector_count": 0,
            })
            logger.info(f"[rebuild:{hospital_id}] 활성 문서 없음 - FAISS 인덱스 초기화")
            return

        _rebuild_status[hospital_id]["progress"] = f"{len(documents)}개 문서 청크 임베딩 중..."

        # 3. 임베딩 모델 설정
        from ....core.config import settings as app_settings
        openai_api_key = os.environ.get("OPENAI_API_KEY") or app_settings.openai_api_key
        if not openai_api_key or openai_api_key == "dev-openai-key":
            _rebuild_status[hospital_id].update({
                "status": "failed",
                "error": "OPENAI_API_KEY가 설정되지 않았습니다.",
                "finished_at": datetime.now().isoformat(),
            })
            return

        embed_model = OpenAIEmbedding(model=EMBEDDING_MODEL, api_key=openai_api_key)
        Settings.embed_model = embed_model

        # 4. FAISS 인덱스 생성
        _rebuild_status[hospital_id]["progress"] = "FAISS 인덱스 생성 중..."
        faiss_index = faiss.IndexFlatL2(EMBEDDING_DIMENSION)
        vector_store = FaissVectorStore(faiss_index=faiss_index)
        storage_context = StorageContext.from_defaults(vector_store=vector_store)

        # 5. 인덱스 빌드 (임베딩 + 저장)
        index = VectorStoreIndex.from_documents(
            documents,
            storage_context=storage_context,
            show_progress=False,
        )

        # 6. 디스크 저장 (rag_service.py가 로드하는 형식과 동일)
        _rebuild_status[hospital_id]["progress"] = "인덱스 저장 중..."
        index.storage_context.persist(persist_dir=str(base))
        # faiss 바이너리 인덱스를 명시적으로 저장 (storage_context.persist는 JSON만 저장)
        faiss.write_index(faiss_index, str(base / "faiss.index"))

        _rebuild_status[hospital_id].update({
            "status": "completed",
            "progress": f"완료: {len(documents)}개 청크, 벡터 {faiss_index.ntotal}개",
            "finished_at": datetime.now().isoformat(),
            "document_count": len(documents),
            "vector_count": faiss_index.ntotal,
        })
        logger.info(f"[rebuild:{hospital_id}] 완료 - 문서 {len(documents)}개, 벡터 {faiss_index.ntotal}개")

    except Exception as e:
        logger.error(f"[rebuild:{hospital_id}] 재구축 실패: {e}")
        logger.error(traceback.format_exc())
        _rebuild_status[hospital_id].update({
            "status": "failed",
            "error": str(e),
            "finished_at": datetime.now().isoformat(),
        })


@router.post("/hospitals/{hospital_id}/rebuild")
async def trigger_rebuild(
    hospital_id: str,
    background_tasks: BackgroundTasks,
    partner_id: str = "welno",
) -> Dict[str, Any]:
    """병원별 임베딩 인덱스 재구축 트리거 (비동기)."""
    # 이미 실행 중이면 중복 방지
    current = _rebuild_status.get(hospital_id)
    if current and current.get("status") == "running":
        return {
            "success": False,
            "message": "이미 재구축이 진행 중입니다.",
            "hospital_id": hospital_id,
            "status": current,
        }
    background_tasks.add_task(_run_rebuild_for_hospital, hospital_id, partner_id)
    return {
        "success": True,
        "message": "재구축이 백그라운드에서 시작되었습니다.",
        "hospital_id": hospital_id,
    }


@router.get("/hospitals/{hospital_id}/rebuild/status")
async def get_rebuild_status(hospital_id: str) -> Dict[str, Any]:
    """재구축 진행 상태 조회"""
    status = _rebuild_status.get(hospital_id)
    if not status:
        return {"hospital_id": hospital_id, "status": "idle", "message": "재구축 이력 없음"}
    return {"hospital_id": hospital_id, **status}


# ─── 병원 RAG 설정 CRUD ──────────────────────────────────────────

@router.get("/hospitals/{hospital_id}/config", response_model=HospitalRagConfigResponse)
async def get_hospital_config(hospital_id: str, partner_id: str = "welno"):
    """병원별 RAG/LLM 설정 조회"""
    try:
        query = """
            SELECT partner_id, hospital_id, hospital_name, contact_phone, persona_prompt, welcome_message,
                   llm_config, embedding_config, theme_config, is_active,
                   created_at, updated_at
            FROM welno.tb_hospital_rag_config
            WHERE partner_id = %s AND hospital_id = %s
        """
        config = await db_manager.execute_one(query, (partner_id, hospital_id))

        if not config:
            hospital_name_query = "SELECT hospital_name FROM welno.welno_hospitals WHERE hospital_id = %s"
            hospital_info = await db_manager.execute_one(hospital_name_query, (hospital_id,))
            hospital_name = hospital_info['hospital_name'] if hospital_info else hospital_id

            return HospitalRagConfigResponse(
                partner_id=partner_id,
                hospital_id=hospital_id,
                hospital_name=hospital_name,
                contact_phone=None,
                persona_prompt="",
                welcome_message="",
                llm_config={"model": "gemini-3-flash-preview", "temperature": 0.7, "max_tokens": 2000},
                embedding_config={"model": "text-embedding-ada-002", "index_name": "faiss_db"},
                theme_config={"theme": "default", "logo_url": None, "primary_color": "#7B5E4F"},
                is_active=True,
            )

        return HospitalRagConfigResponse(
            partner_id=config['partner_id'],
            hospital_id=config['hospital_id'],
            hospital_name=config['hospital_name'],
            contact_phone=config.get('contact_phone'),
            persona_prompt=config['persona_prompt'],
            welcome_message=config['welcome_message'],
            llm_config=config['llm_config'] or {},
            embedding_config=config['embedding_config'] or {},
            theme_config=config['theme_config'] or {},
            is_active=config['is_active'],
            created_at=config['created_at'].isoformat() if config['created_at'] else None,
            updated_at=config['updated_at'].isoformat() if config['updated_at'] else None,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"설정 조회 실패: {str(e)}")


@router.put("/hospitals/{hospital_id}/config")
async def update_hospital_config(
    hospital_id: str,
    config: HospitalRagConfigUpdate = Body(...),
):
    """병원별 RAG/LLM 설정 저장 (Upsert)"""
    try:
        query = """
            INSERT INTO welno.tb_hospital_rag_config
            (partner_id, hospital_id, hospital_name, contact_phone, persona_prompt, welcome_message, llm_config, embedding_config, theme_config, is_active)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (partner_id, hospital_id) DO UPDATE SET
                hospital_name = EXCLUDED.hospital_name,
                contact_phone = EXCLUDED.contact_phone,
                persona_prompt = EXCLUDED.persona_prompt,
                welcome_message = EXCLUDED.welcome_message,
                llm_config = EXCLUDED.llm_config,
                embedding_config = EXCLUDED.embedding_config,
                theme_config = EXCLUDED.theme_config,
                is_active = EXCLUDED.is_active,
                updated_at = NOW()
        """
        await db_manager.execute_update(query, (
            config.partner_id, hospital_id, config.hospital_name, config.contact_phone,
            config.persona_prompt, config.welcome_message,
            json.dumps(config.llm_config or {}), json.dumps(config.embedding_config or {}),
            json.dumps(config.theme_config or {}), config.is_active,
        ))

        from ....services.dynamic_config_service import dynamic_config
        dynamic_config.clear_cache()

        return {"success": True, "hospital_id": hospital_id, "partner_id": config.partner_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"설정 저장 실패: {str(e)}")


@router.post("/hospitals")
async def create_hospital_config(
    config: HospitalRagConfigUpdate = Body(...),
    hospital_id: Optional[str] = None,
):
    """새 병원 RAG 설정 생성"""
    try:
        if not hospital_id:
            import uuid
            hospital_id = str(uuid.uuid4()).replace('-', '').upper()[:32]

        existing = await db_manager.execute_one(
            "SELECT hospital_id FROM welno.tb_hospital_rag_config WHERE partner_id = %s AND hospital_id = %s",
            (config.partner_id, hospital_id),
        )

        if existing:
            raise HTTPException(status_code=409, detail=f"병원 설정이 이미 존재합니다: {hospital_id}")

        query = """
            INSERT INTO welno.tb_hospital_rag_config
            (partner_id, hospital_id, hospital_name, contact_phone, persona_prompt, welcome_message, llm_config, embedding_config, theme_config, is_active)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        await db_manager.execute_update(query, (
            config.partner_id, hospital_id, config.hospital_name, config.contact_phone,
            config.persona_prompt, config.welcome_message,
            json.dumps(config.llm_config or {"model": "gemini-3-flash-preview", "temperature": 0.7, "max_tokens": 2000}),
            json.dumps(config.embedding_config or {"model": "text-embedding-ada-002", "index_name": "faiss_db"}),
            json.dumps(config.theme_config or {"theme": "default", "logo_url": None, "primary_color": "#7B5E4F"}),
            config.is_active,
        ))

        from ....services.dynamic_config_service import dynamic_config
        dynamic_config.clear_cache()

        return {"success": True, "hospital_id": hospital_id, "partner_id": config.partner_id}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"병원 설정 생성 실패: {str(e)}")


@router.delete("/hospitals/{hospital_id}/config")
async def delete_hospital_config(hospital_id: str, partner_id: str = "welno"):
    """병원 RAG 설정 삭제 (비활성화)"""
    try:
        query = """
            UPDATE welno.tb_hospital_rag_config
            SET is_active = false, updated_at = NOW()
            WHERE partner_id = %s AND hospital_id = %s
        """
        result = await db_manager.execute_update(query, (partner_id, hospital_id))

        if result == 0:
            raise HTTPException(status_code=404, detail=f"병원 설정을 찾을 수 없습니다: {hospital_id}")

        from ....services.dynamic_config_service import dynamic_config
        dynamic_config.clear_cache()

        return {"success": True, "hospital_id": hospital_id, "message": "병원 설정이 비활성화되었습니다"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"병원 설정 삭제 실패: {str(e)}")



# ─── 관리자 RAG 테스트 채팅 ────────────────────────────────────────

class TestChatRequest(BaseModel):
    message: str
    partner_id: str = "welno"
    session_id: Optional[str] = None

@router.post("/hospitals/{hospital_id}/test-chat")
async def admin_test_chat(hospital_id: str, body: TestChatRequest):
    """관리자용 RAG 테스트 채팅 (SSE 스트리밍). API Key 인증 없이 동작."""
    import uuid as uuid_mod
    from fastapi.responses import StreamingResponse
    from ....services.partner_rag_chat_service import PartnerRagChatService
    from ....middleware.partner_auth import PartnerAuthInfo

    session_id = body.session_id or f"admin-test-{uuid_mod.uuid4().hex[:8]}"
    
    # 관리자용 mock PartnerAuthInfo
    mock_partner = PartnerAuthInfo(
        partner_id=body.partner_id,
        partner_name="admin-test",
        config={"api_key": "admin-test-key"}
    )
    
    service = PartnerRagChatService()
    
    return StreamingResponse(
        service.handle_partner_message_stream(
            partner_info=mock_partner,
            uuid=f"admin-{uuid_mod.uuid4().hex[:8]}",
            hospital_id=hospital_id,
            message=body.message,
            session_id=session_id,
            partner_health_data={"name": "관리자 테스트", "source": "admin-test"}
        ),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        }
    )
